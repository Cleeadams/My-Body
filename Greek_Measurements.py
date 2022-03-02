import xlsxwriter
from win32com.client import Dispatch
from datetime import date
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from os.path import exists


class Measurements:
    header = ["Date", "Weight", "Neck", "Chest", "Bicep", "Forearm", "Waist", "Hip", "Thigh", "Calve"]

    # inches
    goal = ["in. & lbs", 155, 16, 44, 16, 13, 31, 37, 23, 15]
    # wrist = 6.75
    # chest = 39
    # waist = 31
    # bicep = 13.75
    # neck = 15
    # thigh = 21
    # forearm = 11
    # calve = 15
    # hip = 32.25
    @staticmethod
    def speak(text):
        from gtts import gTTS
        import os
        Speech = gTTS(text=text, lang='en', slow=False)
        Speech.save("text.mp3")
        os.system("start text.mp3")

    # @staticmethod
    # def speech_to_recogn():
    #     import speech_recognition as sr
    #


File = "Greek Measurements.xlsx"
FilePath = "C:\\Users\\conno\\Desktop\\My_PYTHON_Projects\\" + File

x1 = Dispatch("Excel.Application")

if x1.Workbooks.Count > 0:
    if any(i.Name == File for i in x1.Workbooks):
        x1.Workbooks.Close()
        x1.Quit()

if exists(FilePath) == False:
    # Workbook creates
    workbook = xlsxwriter.Workbook('Greek Measurements.xlsx')
    # Worksheets
    worksheet = workbook.add_worksheet('Measurements')
    # Customize cell formats
    worksheet.set_column(0, 0, 15)
    center = workbook.add_format({'align': 'center'})
    format_header = workbook.add_format({'align': 'center', 'bold': True})
    format_goal = workbook.add_format({'align': 'center', 'bold': True, 'font_color': 'blue'})
    # Create Header
    Col = ord('A')
    for x in Measurements.header:
        worksheet.write(chr(Col) + '1', x, format_header)
        Col = Col + 1
    # Create goal row below header
    Col = ord('A')
    for i in Measurements.goal:
        worksheet.write(chr(Col) + '2', i, format_goal)
        Col = Col + 1

    workbook.close()

# Read the Excel file
f = pd.read_excel(FilePath)
row = len(f.axes[0])
book = load_workbook(File)
writer = pd.ExcelWriter(File, engine='openpyxl')
writer.book = book
ws = book['Measurements']

# Data collected
WeightText = "What is your weight for today? "
ChestText = "What is your chest size for today? "
WaistText = "What is your waist size for today? "
BicepText = "What is your bicep size for today? "
NeckText = "What is your neck size for today? "
ThighText = "What is your thigh size for today? "
ForearmText = "What is your forearm size for today? "
CalveText = "What is your calve size for today? "
HipText = "What is your hip size for today? "

date = date.today()
date = str(date.month) + '/' + str(date.day) + '/' + str(date.year)
Measurements.speak(WeightText)
weight = input(WeightText)
Measurements.speak(ChestText)
chest = input(ChestText)
Measurements.speak(WaistText)
waist = input(WaistText)
Measurements.speak(BicepText)
bicep = input(BicepText)
Measurements.speak(NeckText)
neck = input(NeckText)
Measurements.speak(ThighText)
thigh = input(ThighText)
Measurements.speak(ForearmText)
forearm = input(ForearmText)
Measurements.speak(CalveText)
calve = input(CalveText)
Measurements.speak(HipText)
hip = input(HipText)

# weight = 'N/A'
# chest = 39
# waist = 31
# bicep = 13.75
# neck = 15
# thigh = 21
# forearm = 11
# calve = 15
# hip = 32.25

data = [date, weight, neck, chest, bicep, forearm, waist, hip, thigh, calve]
alignment = Alignment(horizontal='center')
col = ord('A')
if date == ws['A' + str(row+1)].value:
    for x in data:
        ws[chr(col) + str(row + 1)] = x
        CC = ws[chr(col) + str(row + 1)]
        CC.alignment = alignment
        col += 1
else:
    for x in data:
        ws[chr(col) + str(row+2)] = x
        CC = ws[chr(col) + str(row+2)]
        CC.alignment = alignment
        col += 1

writer.save()

x1.Visible = True
x1.Workbooks.Open(FilePath)